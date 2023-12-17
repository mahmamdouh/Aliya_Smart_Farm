import RPi.GPIO as GPIO
import dht11
import time
import datetime
from datetime import date
## to be removed
from openpyxl import load_workbook
import numpy  # Import numpy
import pandas as pd


# create sheet
# Load the workbook and select the sheet
#book = load_workbook('new.xlsx')
book = load_workbook('weather.xlsx')
writer = pd.ExcelWriter("excel.xlsx", engine='openpyxl')

writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}
sheet = book['Sheet1']

df_1 = pd.DataFrame(
    [[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]],
    columns=['temp1','Hunid 2','temp2','Hunid 2','temp3','Hunid 3','temp4','Hunid 4','temp5','Hunid 5','temp6','Hunid 6','Day','Time','over all temp','over all humid'])
 
# initialize GPIO
GPIO.setwarnings(True)
GPIO.setmode(GPIO.BCM)

# read data using pin 14
sensor1 = dht11.DHT11(pin=4)
sensor2 = dht11.DHT11(pin=27)
sensor3 = dht11.DHT11(pin=22)
sensor4 = dht11.DHT11(pin=23)
sensor5 = dht11.DHT11(pin=24)
sensor6 = dht11.DHT11(pin=25)

# Diag data Base
DIAG = {
  "DHT_1": False,
  "DHT_2": False,
  "DHT_3": False,
  "DHT_4": False,
  "DHT_5": False,
  "DHT_6": False,
  "DHT_1_Time": "NO LOG",
  "DHT_2_Time": "NO LOG",
  "DHT_3_Time": "NO LOG",
  "DHT_4_Time": "NO LOG",
  "DHT_5_Time": "NO LOG",
  "DHT_6_Time": "NO LOG",
  "DHT_1_Log_Frq": 0,
  "DHT_2_Log_Frq": 0,
  "DHT_3_Log_Frq": 0,
  "DHT_4_Log_Frq": 0,
  "DHT_5_Log_Frq": 0,
  "DHT_6_Log_Frq": 0,
}


# Over all Sensor Data
DHT_OverAll_Temp  = 0
DHT_OverAll_humid = 0
DHT_OverAll_Date  = ""



def DHT_Update_10_min():
    time.sleep(1)
    result1 = sensor1.read()
    result2 = sensor2.read()
    result3 = sensor3.read()
    result4 = sensor4.read()
    result5 = sensor5.read()
    result6 = sensor6.read()
    time.sleep(1)

    # take time stamp if current reading
    ##today = date.today()
    ##now = datetime.datetime.now().time()


    # data from sensor 1
    for x in range(0, 4):
        result1 = sensor1.read()
        if result1.is_valid():
            #DHT_DIAG_Update(1,True)
            DIAG["DHT_1"] = True
            print("getting sensor 1 readings")
            print("Last valid input from sensor 1: " + str(datetime.datetime.now()))
            print("Temperature: %-3.1f C" % result1.temperature)
            print("Humidity: %-3.1f %%" % result1.humidity)
            DHT_DIAG_Update(1, True)
            break;

        else:
            DHT_DIAG_Update(1, False)
            DIAG["DHT_1_Time"] = str(datetime.datetime.now())
            #print(" sensor 1 FAIL to read")

    # data from sensor 2
    for x in range(0, 4):
        result2 = sensor2.read()
        if result2.is_valid():
            #DHT_DIAG_Update(2, True)
            DIAG["DHT_2"] = True
            print("getting sensor 2 readings")
            print("Last valid input from sensor 2: " + str(datetime.datetime.now()))
            print("Temperature: %-3.1f C" % result2.temperature)
            print("Humidity: %-3.1f %%" % result2.humidity)
            DHT_DIAG_Update(2, True)
            break;
        else:
            DHT_DIAG_Update(2, False)
            DIAG["DHT_2_Time"] = str(datetime.datetime.now())
            #print(" sensor 2 FAIL to read")

    # data from sensor 3
    for x in range(0, 4):
        result3 = sensor3.read()
        if result3.is_valid():
            #DHT_DIAG_Update(3, True)
            DIAG["DHT_3"] = True
            print("getting sensor 3 readings")
            print("Last valid input from sensor 3: " + str(datetime.datetime.now()))
            print("Temperature: %-3.1f C" % result3.temperature)
            print("Humidity: %-3.1f %%" % result3.humidity)
            DHT_DIAG_Update(3, True)
            break;

        else:
            DHT_DIAG_Update(3, False)
            DIAG["DHT_3_Time"] = str(datetime.datetime.now())
            #print(" sensor 3 FAIL to read")

    # data from sensor 4
    for x in range(0, 4):
        result4 = sensor4.read()
        if result4.is_valid():
            #DHT_DIAG_Update(4, True)
            DIAG["DHT_4"] = True
            print("getting sensor 4 readings")
            print("Last valid input from sensor 4: " + str(datetime.datetime.now()))
            print("Temperature: %-3.1f C" % result4.temperature)
            print("Humidity: %-3.1f %%" % result4.humidity)
            DHT_DIAG_Update(4, True)
            break;
        else:
            DHT_DIAG_Update(4, False)
            DIAG["DHT_4_Time"] = str(datetime.datetime.now())
            #print(" sensor 4 FAIL to read")

    # data from sensor 5
    for x in range(0, 4):
        result5 = sensor5.read()
        if result5.is_valid():
            #DHT_DIAG_Update(5, True)
            DIAG["DHT_5"] = True
            print("getting sensor 5 readings")
            print("Last valid input from sensor 5: " + str(datetime.datetime.now()))
            print("Temperature: %-3.1f C" % result5.temperature)
            print("Humidity: %-3.1f %%" % result5.humidity)
            DHT_DIAG_Update(5, True)
            break;
        else :
            DHT_DIAG_Update(5, False)
            DIAG["DHT_5_Time"] = str(datetime.datetime.now())
            #print(" sensor 5 FAIL to read")

    # data from sensor 6
    for x in range(0, 4):
        result6 = sensor6.read()
        if result6.is_valid():
            #DHT_DIAG_Update(6, True)
            DIAG["DHT_6"] = True
            print("getting sensor 6 readings")
            print("Last valid input from sensor 6: " + str(datetime.datetime.now()))
            print("Temperature: %-3.1f C" % result6.temperature)
            print("Humidity: %-3.1f %%" % result6.humidity)
            DHT_DIAG_Update(6, True)
            break;
        else:
            DHT_DIAG_Update(6, False)
            DIAG["DHT_6_Time"] = str(datetime.datetime.now())
            #print(" sensor 6 FAIL to read")
    ############################################################################
    today = date.today()
    now = datetime.datetime.now().time()
    Temp , Humid = DHT_Algo_calculator(result1,result2,result3,result4,result5,result6)
    # save data to xcel to be removed
    row = (result1.temperature, result1.humidity,result2.temperature, result2.humidity,result3.temperature, result3.humidity,result4.temperature, result4.humidity,result5.temperature, result5.humidity,result6.temperature, result6.humidity,today, now , Temp , Humid)
    sheet.append(row)
    # using panda data

    df_2 = pd.DataFrame(
        [[result1.temperature, result1.humidity,result2.temperature, result2.humidity,result3.temperature, result3.humidity,result4.temperature, result4.humidity,result5.temperature, result5.humidity,result6.temperature, result6.humidity,today, now , Temp , Humid]],
        columns=['temp1','Hunid 2','temp2','Hunid 2','temp3','Hunid 3','temp4','Hunid 4','temp5','Hunid 5','temp6','Hunid 6','Day','Time','over all temp','over all humid'])

    df_2.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)

    writer.save()
    #print('DataFrame is written successfully to Excel File.')
    
    ############################################################################    
    #print (row)
    # The workbook is saved!
    book.save('weather.xlsx')
    #print ("over all data ")
    #print (Temp , Humid)
    return (Temp , Humid)

# function calculate average of data sensor
def DHT_Algo_calculator(result1,result2,result3,result4,result5,result6):
    average = 0
    average=DHT_Get_DIAG_Status()
    #print ("avrage value is ")
    #print (average)
    if(average == 0):
        return (0,0)
    

    DHT_OverAll_Temp = (result1.temperature + result2.temperature + result3.temperature +result4.temperature + result5.temperature + result6.temperature) / average
    DHT_OverAll_humid = (result1.humidity + result2.humidity + result3.humidity +result4.humidity + result5.humidity + result6.humidity) / average
    #print("Last valid input from overall: " + str(datetime.datetime.now()))
    return (DHT_OverAll_Temp ,DHT_OverAll_humid)




# function update Diag of DHT sensors
def DHT_DIAG_Update(Diag_num,Diag_status):
    ##today = date.today()
    ##now = datetime.datetime.now().time()

    if (Diag_num == 1):
        # update DHT_1 status
        DIAG["DHT_1"] = Diag_status
        #print ("DHT_DIAG_1")
        #print (DIAG["DHT_1"])
        

    if (Diag_num == 2):
        # update DHT_2 status
        DIAG["DHT_2"] = Diag_status
        #print ("DHT_DIAG_2")
        #print (DIAG["DHT_2"])

    if (Diag_num == 3):
        # update DHT_3 status
        DIAG["DHT_3"] = Diag_status
        #print ("DHT_DIAG_3")
        #print (DIAG["DHT_3"])

    if (Diag_num == 4):
        # update DHT_4 status
        DIAG["DHT_4"] = Diag_status
        #print ("DHT_DIAG_4")
        #print ( DIAG["DHT_4"])

    if (Diag_num == 5):
        # update DHT_5 status
        DIAG["DHT_5"] = Diag_status
        #print ("DHT_DIAG_5")
        #print (DIAG["DHT_5"])

    if (Diag_num == 6):
        # update DHT_6 status
        DIAG["DHT_6"] = Diag_status
        #print ("DHT_DIAG_6")
        #print ( DIAG["DHT_6"])



#function to GET Diag status
def DHT_Get_DIAG_Status():
    #umber of workign sensors
    num_of_w_sensor =0
    if (DIAG["DHT_1"] == True):
        num_of_w_sensor=num_of_w_sensor+1
        #print ("DHT_Sensor 1 OK")

    if (DIAG["DHT_2"] == True):
        num_of_w_sensor=num_of_w_sensor+1
        #print ("DHT_Sensor 2 OK")

    if (DIAG["DHT_3"] == True):
        num_of_w_sensor=num_of_w_sensor+1
        #print ("DHT_Sensor 3 OK")

    if (DIAG["DHT_4"] == True):
        num_of_w_sensor=num_of_w_sensor+1
        #print ("DHT_Sensor 4 OK")

    if (DIAG["DHT_5"] == True):
        num_of_w_sensor=num_of_w_sensor+1
        #print ("DHT_Sensor 5 OK")

    if (DIAG["DHT_6"] == True):
        num_of_w_sensor=num_of_w_sensor+1
        #print ("DHT_Sensor 6 OK")
        
    return (num_of_w_sensor)


#function to GET Diag status
def DHT_Get_DIAG():
    #umber of workign sensors
    if (DIAG["DHT_1"] == False):
        print ("DHT_Sensor 1 have error")
        print ("DHT_Sensor 1 error time ")
        print (DIAG["DHT_1_Time"])
    else:
        print ("DHT_Sensor 1 OK")

    if (DIAG["DHT_2"] == False):
        print ("DHT_Sensor 2 have error")
        print ("DHT_Sensor 2 error time")
        print (DIAG["DHT_2_Time"])
    else:
        print ("DHT_Sensor 2 OK")

    if (DIAG["DHT_3"] == False):
        print ("DHT_Sensor 3 have error")
        print ("DHT_Sensor 3 error time ")
        print (DIAG["DHT_3_Time"])
    else:
        print ("DHT_Sensor 3 OK")

    if (DIAG["DHT_4"] == False):
        print ("DHT_Sensor 4 have error")
        print ("DHT_Sensor 4 error time  ")
        print (DIAG["DHT_4_Time"])
    else:
        print ("DHT_Sensor 4 OK")

    if (DIAG["DHT_5"] == False):
        print ("DHT_Sensor 5 have error")
        print ("DHT_Sensor 5 error time ")
        print (DIAG["DHT_5_Time"])
    else:
        print ("DHT_Sensor 5 OK")

    if (DIAG["DHT_6"] == False):
        print ("DHT_Sensor 6 have error")
        print ("DHT_Sensor 6 error time ")
        print (DIAG["DHT_6_Time"])
    else:
        print ("DHT_Sensor 6 OK")
        



