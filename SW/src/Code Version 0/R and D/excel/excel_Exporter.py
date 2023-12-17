import openpyxl
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import re
import logging
import getpass
import datetime
import os
import pandas as pd
from openpyxl import load_workbook
import pandas as pd

 

df = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})
book = load_workbook('Output_Format.xlsx')
writer = pd.ExcelWriter('Output_Format.xlsx')
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets) # *I am not sure what is happening in this line*
df.to_excel(writer,"Sheet1",startcol=0, startrow=20)
#df.append_df_to_excel('Output_Format.xlsx', df, sheet_name="Sheet1", startcol=0, startrow=20)

#Create the output sheet
def write_to_Excel(time , data , row , cloum):
    wb2 = load_workbook(filename = 'Output_Format.xlsx')
    sheet.cell_value(0, 0)	
    Sheet1 = wb2.sheet_by_index(0)
    Outsheet = wb2.get_sheet_by_name('Sheet1')
    # For row 0 and column 0 

    Outsheet.cell(row,cloum).value = time
    Outsheet.cell(row,cloum+1).value = time
    wb2.save("Output_Format.xlsx")
    wb2.close()

